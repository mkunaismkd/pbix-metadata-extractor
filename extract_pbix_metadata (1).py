#!/usr/bin/env python3
"""
PBIX Metadata Extractor
------------------------
Extracts metadata from a Power BI (.pbix) file - tables, columns/schema,
DAX measures, calculated columns, calculated tables, relationships,
Power Query (M) code, parameters, and row-level security roles - and
exports everything into a single, formatted Excel workbook.

Usage:
    python extract_pbix_metadata.py path/to/file.pbix
    python extract_pbix_metadata.py path/to/file.pbix -o output.xlsx
    python extract_pbix_metadata.py path/to/folder/          # batch mode

Requires: pbixray, pandas, openpyxl
    pip install pbixray pandas openpyxl
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from pbixray import PBIXRay
except ImportError:
    sys.exit("Missing dependency. Install with: pip install pbixray pandas openpyxl")


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def safe_df(getter, *args, **kwargs) -> pd.DataFrame:
    """Call a PBIXRay property/method and always return a DataFrame,
    even if the model has none of that object type or the attribute
    isn't supported by this file's schema version."""
    try:
        result = getter(*args, **kwargs) if callable(getter) else getter
        if result is None:
            return pd.DataFrame()
        df = result if isinstance(result, pd.DataFrame) else pd.DataFrame(result)
        return df
    except Exception as e:
        print(f"  [warn] could not read {getattr(getter, '__name__', getter)}: {e}")
        return pd.DataFrame()


def build_overview(model: PBIXRay, file_path: Path) -> pd.DataFrame:
    meta = safe_df(model.metadata)
    stats = safe_df(model.statistics)

    rows = [
        ("File name", file_path.name),
        ("File size (bytes)", getattr(model, "size", "n/a")),
        ("Table count", len(getattr(model, "tables", []))),
        ("Measure count", len(safe_df(lambda: model.dax_measures))),
        ("Calculated column count", len(safe_df(lambda: model.dax_columns))),
        ("Calculated table count", len(safe_df(lambda: model.dax_tables))),
        ("Relationship count", len(safe_df(lambda: model.relationships))),
        ("Parameter count", len(safe_df(lambda: model.m_parameters))),
    ]
    overview = pd.DataFrame(rows, columns=["Property", "Value"])

    if not meta.empty:
        # PBIXRay already returns metadata as Name/Value pairs
        meta_rows = meta.rename(columns={"Name": "Property"})[["Property", "Value"]]
        overview = pd.concat([overview, meta_rows], ignore_index=True)

    return overview


def build_tables_sheet(model: PBIXRay) -> pd.DataFrame:
    tables = getattr(model, "tables", [])
    return pd.DataFrame({"Table Name": list(tables)})


def extract_all(pbix_path: Path) -> dict:
    print(f"Reading {pbix_path.name} ...")
    model = PBIXRay(str(pbix_path))

    sheets = {
        "Overview": build_overview(model, pbix_path),
        "Tables": build_tables_sheet(model),
        "Columns (Schema)": safe_df(lambda: model.schema),
        "Measures (DAX)": safe_df(lambda: model.dax_measures),
        "Calculated Columns": safe_df(lambda: model.dax_columns),
        "Calculated Tables": safe_df(lambda: model.dax_tables),
        "Relationships": safe_df(lambda: model.relationships),
        "Power Query (M)": safe_df(lambda: model.power_query),
        "Parameters": safe_df(lambda: model.m_parameters),
        "Perspectives": safe_df(lambda: model.perspectives),
        "Row-Level Security": safe_df(lambda: model.rls),
    }

    model.close()
    # Drop sheets that came back completely empty so the workbook stays clean
    return {name: df for name, df in sheets.items() if not (df.empty and name not in ("Overview", "Tables"))}


def autosize_and_style(ws, df: pd.DataFrame):
    ws.freeze_panes = "A2"
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

        # Reasonable column width capped for long DAX/M expressions
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)]
        )
        width = min(max(max_len + 2, 12), 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Wrap long text columns (DAX/M expressions) so they're readable
        if "expression" in str(col_name).lower() or "query" in str(col_name).lower():
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = WRAP_ALIGN


def write_excel(sheets: dict, output_path: Path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = name[:31]  # Excel sheet name limit
            df.to_excel(writer, sheet_name=safe_name, index=False)
            autosize_and_style(writer.sheets[safe_name], df)
    print(f"Saved: {output_path}")


def process_file(pbix_path: Path, output_dir: Path = None):
    sheets = extract_all(pbix_path)
    out_dir = output_dir or pbix_path.parent
    output_path = out_dir / f"{pbix_path.stem}_metadata.xlsx"
    write_excel(sheets, output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Extract Power BI (.pbix) metadata to Excel.")
    parser.add_argument("input", help="Path to a .pbix file, or a folder for batch mode")
    parser.add_argument("-o", "--output", help="Output .xlsx path (single-file mode only)")
    args = parser.parse_args()

    input_path = Path(args.input)

    if input_path.is_dir():
        pbix_files = sorted(input_path.glob("*.pbix"))
        if not pbix_files:
            sys.exit(f"No .pbix files found in {input_path}")
        for f in pbix_files:
            try:
                process_file(f, output_dir=input_path)
            except Exception as e:
                print(f"  [error] failed on {f.name}: {e}")
    elif input_path.is_file():
        output_path = Path(args.output) if args.output else None
        if output_path:
            sheets = extract_all(input_path)
            write_excel(sheets, output_path)
        else:
            process_file(input_path)
    else:
        sys.exit(f"Path not found: {input_path}")


if __name__ == "__main__":
    main()
